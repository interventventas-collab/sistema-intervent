"""Genera un Excel de prueba para la importacion por OEM con productos Colombraro.

Usa la columna 'sku_interno' para que cada fila tenga el SKU correcto que
matchea con la publicacion en Mercado Libre (ej: C9335GR para el armario gris).

NO incluye columna stock — los productos arrancan en 0 y se cargan despues
desde "Modificacion de stock".
"""
import datetime
import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Productos por OEM"

now = datetime.datetime(2026, 4, 30, 12, 0, 0)
wb.properties.created = now
wb.properties.modified = now
wb.properties.creator = "AIml"
wb.properties.lastModifiedBy = "AIml"

headers = [
    "codigo_oem",
    "sku_interno",
    "titulo",
    "marca",
    "precio_costo",
    "precio_venta_con_iva",
    "iva",
    "codigo_de_barras",
    "stock_critico",
]

for col_idx, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = Font(bold=True)

# Productos de prueba con el SKU que matchea ML
# (oem, sku_interno, titulo, marca, costo_sin_iva, pvp_con_iva, iva, barcode, critico)
rows = [
    # Productos con publicacion ML existente -> el sku_interno matchea SellerSku
    ("136",  "C136BL",   "Batea Contenedor Plástico 35x49x6,5 Blanco",          "COLOMBRARO", 4524.00,    9520.00,  21, "7790733001360", 5),
    ("818",  "C818BL",   "Organizador N2 con divisiones movibles - Blanco",     "COLOMBRARO", 3469.50,    7280.00,  21, "7790733008185", 3),
    ("1495", "C1495NEG", "Bandeja para Papeles A4 Apilable - Negra",            "COLOMBRARO", 4400.00,    7999.00,  21, "7790733014957", 5),
    ("8718", "C8718BL",  "Organizadores Interconectables Set x4 - Blanco",      "COLOMBRARO", 4310.25,    8999.00,  21, "7790733087180", 10),
    ("9399", "C9399TR",  "Caja Ideal Zapatos 10 Lts - Transparente",            "COLOMBRARO", 7446.75,   15499.00,  21, "7790733093990", 3),
    # Producto con 2 colores: el OEM 8733 se usa para 2 filas
    ("8733", "C8733BL",  "Cajonera en Torre x 3 Grande - Blanca",               "COLOMBRARO", 35501.25, 79520.00,  21, "7790733087333", 1),
    ("8733", "C8733NEG", "Cajonera en Torre x 3 Grande - Negra",                "COLOMBRARO", 35501.25, 79520.00,  21, "7790733087333", 1),
    # Producto sin variantes ni ML -> sku_interno = OEM (lo dejamos vacio para usar default)
    ("103",  None,        "Balde Europa 11 Lts.",                               "COLOMBRARO", 4618.50,    9410.00,  21, "7790733001032", 0),
]

for r_idx, row in enumerate(rows, start=2):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

widths = [12, 14, 50, 14, 14, 16, 6, 18, 10]
for col_idx, w in enumerate(widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w

out = "/home/dev/ai-ml/test_data/colombraro_oem_test.xlsx"
wb.save(out)

# Workaround: openpyxl escribe 'modified' con formato malformado que ClosedXML rechaza.
import zipfile, shutil, re, os
tmp = out + ".tmp"
shutil.copy(out, tmp)
with zipfile.ZipFile(tmp, "r") as zin, zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
    for item in zin.namelist():
        data = zin.read(item)
        if item == "docProps/core.xml":
            xml = data.decode("utf-8")
            xml = re.sub(
                r"(<dcterms:(?:created|modified)[^>]*>)[^<]+(</dcterms:(?:created|modified)>)",
                r"\g<1>2026-04-30T12:00:00Z\g<2>",
                xml,
            )
            data = xml.encode("utf-8")
        zout.writestr(item, data)
os.remove(tmp)

print(f"Excel creado: {out}")
print(f"Filas de datos: {len(rows)}")
